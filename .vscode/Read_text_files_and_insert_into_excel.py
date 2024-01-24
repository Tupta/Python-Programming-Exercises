# Program odczytujący zawartość kilku plików
# tekstowych i wstawiający # ją do arkusza 
# kalkulacyjnego, po jednym wierszu tekstu w
# każdym wierszu arkusza. Wiersze pierwszego pliku
# tekstowego będą komórkami kolumny A, drugiego pliku
# tekstowego będą komórkami kolumny B i tak dalej.
# Wykorzystaj metodę readlines() obiektu File do otrzymania
# listy ciągów tekstowych, po jednym ciągu tekstowym dla
# każdego wiersza w pliku tekstowym. W pierwszym pliku
# pierwszy wiersz tekstu powinien znaleźć się w kolumnie
# pierwszej wiersza pierwszego. Drugi wiersz tekstu powinien
# znaleźć się w kolumnie pierwszej wiersza drugiego i tak
# dalej. Tekst odczytany za pomocą readlines() z pliku drugiego
# powinien trafić do kolumny drugiej, z pliku trzeciego do
# kolumny trzeciej i tak dalej.

import openpyxl

def read_text_files_and_insert_into_excel(*file_paths):
    # Utwórz nowy arkusz kalkulacyjny
    wb = openpyxl.Workbook()
    sheet = wb.active

    # Dla każdego pliku tekstowego
    for col_num, file_path in enumerate(file_paths, start=1):
        # Otwórz plik tekstowy
        with open(file_path, 'r') as file:
            # Odczytaj linie tekstu z pliku
            lines = file.readlines()

            # Dla każdej linii tekstu, wstaw do odpowiedniej kolumny w arkuszu
            for row_num, line in enumerate(lines, start=1):
                # Usuń znaki nowej linii na końcu linii
                line = line.rstrip('\n')

                # Wstaw do arkusza
                sheet.cell(row=row_num, column=col_num, value=line)

    # Zapisz arkusz kalkulacyjny do pliku
    wb.save('output.xlsx')

if __name__ == "__main__":
    # Przykładowe ścieżki do plików tekstowych
    file1_path = 'file1.txt'
    file2_path = 'file2.txt'
    file3_path = 'file3.txt'

    # Wywołaj funkcję z podanymi ścieżkami do plików tekstowych
    read_text_files_and_insert_into_excel(file1_path, file2_path, file3_path)

#################################################
###################     ponizej 2 wersja

#################################################

import openpyxl

def wczytaj_pliki_do_arkusza(*nazwy_plikow, nazwa_arkusza='Arkusz1'):
    wb = openpyxl.Workbook()
    arkusz = wb.active
    arkusz.title = nazwa_arkusza

    for i, nazwa_pliku in enumerate(nazwy_plikow):
        with open(nazwa_pliku, 'r', encoding='utf-8') as plik:
            zawartosc = plik.readlines()
            for wiersz, tekst in enumerate(zawartosc, start=1):
                komorka = arkusz.cell(row=wiersz, column=i+1, value=tekst.strip())

    wb.save('wynik.xlsx')
    print(f'Pliki wczytane do arkusza. Wynik zapisano w pliku "wynik.xlsx".')

# Przykładowe użycie funkcji
wczytaj_pliki_do_arkusza('plik1.txt', 'plik2.txt', 'plik3.txt')
