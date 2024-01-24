#UpdateProduce
# 
# Projekt — uaktualnienie skoroszytu

#! python3

import openpyxl

wb = openpyxl.load_workbook('C:\\Users\\sendrat\\Documents\\python\\.vscode\\dodatkowe materialy\\automatyzacja_pliki\\produceSales.xlsx')
sheet = wb.get_sheet_by_name('Sheet')

PRICE_UPDATES = { 'Garlic' : 3.07,
                 'Celery' : 1.19,
                 'Lemon' :1.27
                 }

# TODO: Iteracja przez wiersze i uaktualnienie cen.

for rowNum in range(2, sheet.max_row):
    produceName = sheet.cell(row=rowNum, column = 1).value
    if produceName in PRICE_UPDATES:
        sheet.cell(row = rowNum, column=2).value = PRICE_UPDATES[produceName]
        wb.save =('UpdateProduce02.xlsx')