# import openpyxl

# from openpyxl.utils import *

# # from openpyxl import load_workbook
# # from openpyxl.utils import get_column_letter

# wb = openpyxl.load_workbook('C:\\Users\\sendrat\\Documents\\python\\.vscode\\example.xlsx')
# sheet = wb['Arkusz1']
# max_row = sheet.max_row
# max_high = sheet.max_column
# print(f"Najwyższy numer wiersza: {max_row} stefan {max_high}")

# jj = get_column_letter(sheet.max_column)
# print(jj)

########################################################################################################
#########################################################################################################

import openpyxl

wb = openpyxl.Workbook()
wb.get_sheet_names()
sheet = wb.active
sheet.title = 'Spam Bacon Eggs Sheet'
print(wb.get_sheet_names())

wb = openpyxl.load_workbook('C:\\Users\\sendrat\\Documents\\python\\.vscode\\example.xlsx')
sheet = wb.active
sheet.title = 'Spam'
wb.save('example_copy.xlsx') 
wb.create_sheet(index=0, title='Pierwszy nowy arkusz')
wb.create_sheet(index=1, title='drugi nowy arkusz')
wb.create_sheet(index=2, title='trzeci nowy arkusz')
print(wb.get_sheet_names())
wb.remove_sheet(wb.get_sheet_by_name('drugi nowy arkusz'))
print(wb.get_sheet_names())
sheet = wb.get_sheet_by_name('Spam')
sheet['A1'] = 'Witaj, świecie!'
print(sheet['A1'].value)