# import openpyxl
# from openpyxl.chart import BarChart, Reference

# wb = openpyxl.Workbook()
# #sheet = wb.active
# sheet = wb.get_sheet_by_name('Sheet')
# for i in range(1, 11):# Wygenerowanie pewnych danych w kolumnie A.
#     sheet['A' + str(i)] = i
# refObj = openpyxl.charts.Reference(sheet, (1, 1), (10, 1))
# seriesObj = openpyxl.charts.Series(refObj, title='Pierwsza seria')
# chartObj = openpyxl.charts.BarChart()
# chartObj.append(seriesObj)
# chartObj.drawing.top = 50 # Zdefiniowanie położenia wykresu.
# chartObj.drawing.left = 100
# chartObj.drawing.width = 370 # Określenie wielkości wykresu.
# chartObj.drawing.height = 200
# sheet.add_chart(chartObj)
# wb.save('sampleChart.xlsx')

##########################################################################################2 wersjA
# import openpyxl
# from openpyxl.chart import BarChart, Reference

# wb = openpyxl.Workbook()
# sheet = wb.active

# for i in range(1, 11):
#     sheet['A' + str(i)] = i

# refObj = Reference(sheet, min_col=1, min_row=1, max_col=1, max_row=10)
# seriesObj = openpyxl.chart.Series(refObj, title='Pierwsza seria')
# chartObj = BarChart()
# chartObj.append(seriesObj)
# chartObj.drawing.top = 50
# chartObj.drawing.left = 100
# chartObj.drawing.width = 370
# chartObj.drawing.height = 200
# sheet.add_chart(chartObj)

# wb.save('sampleChart.xlsx')

############################################################ 3 wersja


import openpyxl, os
from openpyxl.chart import BarChart, Reference

wb = openpyxl.Workbook()
sheet = wb.active

for i in range(1, 11):
    sheet['A' + str(i)] = i

refObj = Reference(sheet, min_col=1, min_row=1, max_col=1, max_row=10)
seriesObj = openpyxl.chart.Series(refObj, title='Pierwsza seria')
chartObj = BarChart()
chartObj.append(seriesObj)

# Ustawienie pozycji wykresu bezpośrednio na obiekcie chartObj
chartObj.left = 100
chartObj.top = 50

# Określenie wielkości wykresu
chartObj.width = 370
chartObj.height = 200

sheet.add_chart(chartObj)
# Wyświetlenie aktualnej ścieżki zapisu

print(os.path.abspath('sampleChart.xlsx'))
wb.save('C:\\Users\\sendrat\\Documents\\python\\.vscode\\sampleChart.xlsx')
